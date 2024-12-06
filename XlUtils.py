import openpyxl

cached_data = {}


def get_row_count(file_path, sheet_name):
    """
    Counts the number of rows in the Excel sheet
    :param file_path: Path of the Excel file
    :param sheet_name: Excel sheet to be read
    :return: int : Number of rows in the Excel sheet including the header
    """
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.get_sheet_by_name(sheet_name)

    # sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file_path, sheet_name):
    """
    Counts the number of columns in the Excel sheet
    :param file_path: Path of the Excel file
    :param sheet_name: Excel sheet to be read
    :return: int : Number of columns in the Excel sheet
    """
    workbook = openpyxl.load_workbook(file_path)

    # sheet = workbook[sheet_name]

    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_column


def get_cell_data(file_path, sheet_name, row, column):
    """
    Reads a particular cell in the Excel sheet
    :param file_path: Path of the Excel file
    :param sheet_name: Excel sheet to be read
    :param row: Row to be read
    :param column: Column to be read
    :return: string : The value in the requested cell
    """
    workbook = openpyxl.load_workbook(file_path)

    # sheet = workbook[sheet_name]
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row=row, column=column).value


def set_cell_data(file_path, sheet_name, row, column, data):
    """
    Writes to a particular cell in the excel sheet
    :param file_path: Path of the Excel file
    :param sheet_name: Excel sheet to be read
    :param row: Row to be read
    :param column: Column to be read
    :param data: Data to be written in the cell specified by row and column
    :return: bool: True if the write was successful else False
    """
    workbook = openpyxl.load_workbook(file_path)

    # sheet = workbook[sheet_name]
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row=row, column=column).value = data
    return workbook.save(file_path)


def get_data_from_excel(file_path, sheet_name):
    """
    Reads all the data from an Excel sheet
    :param file_path: Path of the Excel file
    :param sheet_name: Excel sheet to be read
    :return: list of lists : Data from the Excel sheet
    """
    data = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    row_count = sheet.max_row
    col_count = sheet.max_column

    for row in range(2, row_count + 1):
        rows = []
        for col in range(1, col_count + 1):
            rows.append(sheet.cell(row=row, column=col).value)
        data.append(rows)

    return data
