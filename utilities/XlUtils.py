import openpyxl


def get_row_count(file_path, sheet_name):
    """
    Counts the number of rows in the Excel sheet
    :param file_path: Path of the Excel file
    :param sheet_name: Excel sheet to be read
    :return: int : Number of rows in the Excel sheet including the header
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.max_row
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return None
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found in file '{file_path}'.")
        return None


def get_column_count(file_path, sheet_name):
    """
    Counts the number of columns in the Excel sheet
    :param file_path: Path of the Excel file
    :param sheet_name: Excel sheet to be read
    :return: int : Number of columns in the Excel sheet
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.max_column
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return None
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found in file '{file_path}'.")
        return None


def get_cell_data(file_path, sheet_name, row, column):
    """
    Reads a particular cell in the Excel sheet
    :param file_path: Path of the Excel file
    :param sheet_name: Excel sheet to be read
    :param row: Row to be read
    :param column: Column to be read
    :return: string : The value in the requested cell
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row, column=column).value
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        return None
    except KeyError:
        print(f"Error: Sheet '{sheet_name}' not found in file '{file_path}'.")
        return None
